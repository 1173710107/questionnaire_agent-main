import json
import os
from typing import List, Dict, Any


class ExcelQuestionnaireImporter:
    """
    Excel问卷模板导入工具
    支持将Excel模板转换为问卷JSON格式，保留答案用于自动打分
    
    Excel模板格式：
    | 题目类型 | 题干内容 | 答案解析 | 正确答案 | 选项A | 选项B | 选项C | 选项D |
    
    题型支持：单选题、多选题、评分题、判断题
    正确答案格式：
    - 单选题：A/B/C/D
    - 多选题：AB/ACD等（选项字母组合）
    - 评分题：1/2/3/4/5
    - 判断题：正确/错误
    """
    
    def __init__(self):
        self.supported_types = ["单选题", "多选题", "评分题", "判断题"]
    
    def _load_workbook(self, excel_path: str):
        ext = os.path.splitext(excel_path)[1].lower()
        
        if ext == '.xlsx':
            from openpyxl import load_workbook
            return load_workbook(excel_path), 'openpyxl'
        elif ext == '.xls':
            import xlrd
            return xlrd.open_workbook(excel_path, encoding_override='utf-8'), 'xlrd'
        else:
            raise ValueError("不支持的文件格式: %s，仅支持 .xls 和 .xlsx" % ext)
    
    def import_from_excel(self, excel_path: str) -> Dict[str, Any]:
        if not os.path.exists(excel_path):
            raise FileNotFoundError("Excel文件不存在: %s" % excel_path)
        
        wb, lib_type = self._load_workbook(excel_path)
        
        questionnaire = {
            "title": "基层基础提升调查问卷",
            "subtitle": "",
            "description": "从Excel模板导入的问卷",
            "questions": []
        }
        
        questions = []
        
        if lib_type == 'openpyxl':
            sheet = wb.active
            for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_num == 1:
                    continue
                if not row[0]:
                    continue
                try:
                    question = self._parse_row(row, row_num)
                    if question:
                        questions.append(question)
                except Exception as e:
                    print("解析第%d行时出错: %s" % (row_num, e))
        else:
            sheet = wb.sheet_by_index(0)
            for row_num in range(1, sheet.nrows):
                row = sheet.row_values(row_num)
                if not row[0]:
                    continue
                try:
                    question = self._parse_row(row, row_num + 1)
                    if question:
                        questions.append(question)
                except Exception as e:
                    print("解析第%d行时出错: %s" % (row_num + 1, e))
        
        questionnaire["questions"] = questions
        return questionnaire
    
    def _format_value(self, value):
        """将所有值转为字符串，保持原始格式"""
        if isinstance(value, float):
            if value == int(value):
                return str(int(value))
        return str(value)
    
    def _parse_row(self, row, row_num: int) -> Dict[str, Any]:
        """解析单行数据，Excel格式：
        列0: 题目类型, 列1: 题干内容, 列2: 答案解析, 列3: 正确答案
        列4: 选项A, 列5: 选项B, 列6: 选项C, 列7: 选项D, 列8: 选项E
        """
        q_type = str(row[0]).strip() if row[0] else "单选题"
        
        if q_type not in self.supported_types:
            print("第%d行: 不支持的题型 '%s'，已跳过" % (row_num, q_type))
            return None
        
        title = str(row[1]).strip() if row[1] else ""
        if not title:
            print("第%d行: 题目内容为空，已跳过" % row_num)
            return None
        
        correct_answer = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        
        options = []
        option_labels = ["A", "B", "C", "D", "E"]
        for i in range(5):
            col_idx = 4 + i
            if col_idx < len(row) and row[col_idx]:
                opt_value = self._format_value(row[col_idx]).strip()
                if opt_value:
                    options.append(opt_value)
        
        question = {
            "id": row_num - 1,
            "type": q_type,
            "title": title,
            "correct_answer": correct_answer,
            "score": 10
        }
        
        if q_type != "判断题":
            question["options"] = options
        
        return question
    
    def save_to_json(self, questionnaire: Dict[str, Any], output_path: str):
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(questionnaire, f, ensure_ascii=False, indent=4)
        
        print("问卷已保存到: %s" % output_path)
    
    def import_and_save(self, excel_path: str, output_path: str = None) -> str:
        questionnaire = self.import_from_excel(excel_path)
        
        if not output_path:
            base_name = os.path.splitext(os.path.basename(excel_path))[0]
            output_path = "output/%s.json" % base_name
        
        self.save_to_json(questionnaire, output_path)
        return output_path


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='从Excel模板导入问卷')
    parser.add_argument('--input', type=str, required=True, help='Excel模板文件路径')
    parser.add_argument('--output', type=str, help='输出JSON文件路径')
    
    args = parser.parse_args()
    
    importer = ExcelQuestionnaireImporter()
    
    try:
        output_path = importer.import_and_save(args.input, args.output)
        print("成功从Excel导入问卷！")
        print("输入文件: %s" % args.input)
        print("输出文件: %s" % output_path)
        
        with open(output_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            print("\n问卷信息:")
            print("  - 标题: %s" % data['title'])
            print("  - 问题数量: %d" % len(data['questions']))
            
            type_counts = {}
            for q in data['questions']:
                q_type = q['type']
                type_counts[q_type] = type_counts.get(q_type, 0) + 1
            
            print("  - 题型分布: %s" % type_counts)
            
    except Exception as e:
        print("导入失败: %s" % e)


if __name__ == "__main__":
    main()
